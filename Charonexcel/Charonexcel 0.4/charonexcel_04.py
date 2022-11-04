from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.cell import column_index_from_string
import os
import time

category_list_len=0
category_value=0
example_count=1
num_product_lines=0
pro=0
process=0
program_status=True

category_col=' '
command=' '

#===============================================================================================================================================
#БЛОК ФУНКЦИЙ

def numRows_feed(feed_sheet_file): #Функция вычисления общего числа строк в фиде данных
    num_rows=feed_sheet_file.max_row
    return num_rows

def dataCell_coordinates(row_num, cc_n): #Функция составления координаты ячейки
    CC=cc_n+str(row_num)
    return CC

def get_categories_list(feed_sheet_file, category_col): #Функция составления списка категорий всех товаров в фиде
    c_list=[]
    cat_col_num=column_index_from_string(category_col)
    for i in feed_sheet_file.iter_cols(min_col=cat_col_num, max_col=cat_col_num):
        for j in i:
            c_list.append(j.value)
    c_list.pop(0)
    return c_list

def get_unique_categories_list(cat_list): #Функция составления списка уникальных категорий в фиде
    list_of_unique_categories=[]
    unique_cat=set(cat_list)
    for i in unique_cat:
        list_of_unique_categories.append(i)
    return list_of_unique_categories

def get_filling_process(category_list_len, pro): #Функция вывода процесса заполнения
    process=round((pro/category_list_len)*100, 1)
    return process

#===============================================================================================================================================
#ОБЩИЙ ЦИКЛ ПРОГРАММЫ

while program_status==True:

#===============================================================================================================================================
#МЕНЮ ПРОГРАММЫ
    
    print('CHARONEXCEL v. 0.4' + '\n\nВЫБЕРИТЕ КОМАНДУ:\nE - ЗАПУСТИТЬ ПРОГРАММУ\nQ - ЗАКРЫТЬ ПРОГРАММУ')
    command=input('КОМАНДА: ')
    if command=='E' or command=='e':
    
#===============================================================================================================================================
#НАЧАЛО РАБОТЫ
        
        main_time=time.time()
        
#===============================================================================================================================================
#ОБЪЯВЛЕНИЕ ФИДА
        
        feed_file='excels\A.xlsx'
        feed_load_file=load_workbook(feed_file)
        feed_sheet_file=feed_load_file['Sheet1']
        
#===============================================================================================================================================
#ПОИСК СТОЛБЦА КАТЕГОРИЙ
        
        for feed_row_cells in feed_sheet_file.iter_rows(min_row=1, max_row=1):
            for feed_cell in feed_row_cells:
                if feed_cell.value=='Категория' or feed_cell.value=='Category':
                    category_col=feed_cell.column_letter
        
#===============================================================================================================================================
#ПОИСК УНИКАЛЬНЫХ КАТЕГОРИЙ
        
        c_list=get_categories_list(feed_sheet_file, category_col)
        category_list=get_unique_categories_list(c_list)
    #    print(category_list)
        category_list_len=len(category_list)
        
#===============================================================================================================================================
#ЧТЕНИЕ ПАПКИ С ШАБЛОНАМИ
        
        listok=os.listdir("excels")
    #    print(listok, '\n ')
        
        for alpha in category_list:
            alpha_s='(' + alpha + ')'
            for beta in listok:
                if alpha_s in beta:
                    start_time=time.time()
                    category_value=alpha
        
#===============================================================================================================================================
#ОБЪЯВЛЕНИЕ ШАБЛОНА
                    
                    example_file='excels\\' + beta
                    example_load_file=load_workbook(example_file)
                    example_sheet_file=example_load_file['Sheet1']
                    
#===============================================================================================================================================
#ДОБАВЛЕНИЕ ДОПОЛНИТЕЛЬНОЙ ШАПКИ
    
                    for feed_row_cells in feed_sheet_file.iter_rows(min_row=1, max_row=1):
                        for feed_cell in feed_row_cells:
                            if feed_cell.value=='Image11':
                                        min_feedRuInfo=feed_cell.column+1
                                        for rusin in range(min_feedRuInfo, feed_sheet_file.max_column+1):
                                            head_feed_rusin=dataCell_coordinates(1, get_column_letter(rusin))
                                            last_examp_rusin=dataCell_coordinates(1, get_column_letter(example_sheet_file.max_column+1))
                                            example_sheet_file[last_examp_rusin]=feed_sheet_file[head_feed_rusin].value
        
#===============================================================================================================================================
#ПОИСК СТОЛБЦА КАТЕГОРИЙ
                    
                    for feed_row_cells in feed_sheet_file.iter_rows(min_row=1, max_row=1):
                        for feed_cell in feed_row_cells:
                            if feed_cell.value=='Категория' or feed_cell.value=='Category':
                                category_col=feed_cell.column_letter
                                
#===============================================================================================================================================
#ПЕРЕНОС ДАННЫХ С ФИДА В ШАБЛОН
                    
                    feed_NUM_ROWS=numRows_feed(feed_sheet_file)+1
                    for feed_row_cells in feed_sheet_file.iter_rows(min_row=1, max_row=1):
                        for feed_cell in feed_row_cells:
                                        
                            for example_row_cells in example_sheet_file.iter_rows(min_row=1, max_row=1):
                                for example_cell in example_row_cells:
                                    
                                    example_count=1
                                    num_product_lines=0
                                    if feed_cell.value==example_cell.value:                    
                                        E_column_num=example_cell.column_letter
                                        F_column_num=feed_cell.column_letter
                                        
                                        for row_num in range(2, feed_NUM_ROWS):
                                            cat_c=dataCell_coordinates(row_num, category_col)
                                            if(feed_sheet_file[cat_c].value==category_value):
                                                example_count+=1
                                                example_sheet_file[dataCell_coordinates(example_count, E_column_num)]=feed_sheet_file[dataCell_coordinates(row_num, F_column_num)].value
                                                num_product_lines=num_product_lines+1
    
#===============================================================================================================================================
#СОХРАНЕНИЕ ШАБЛОНА
                         
                    example_load_file.save(example_file)
                    example_load_file.close()
                  
#===============================================================================================================================================
#ВЫВОД ПРОЦЕССА ЗАПОЛНЕНИЯ
                  
                    pro=pro+1
                    process=get_filling_process(category_list_len, pro)
                    print('Шаблон ' + beta + ' заполнен')
                    print('КОЛИЧЕСТВО СТРОК ШАБЛОНА:', num_product_lines)
                    print('ВРЕМЯ ЗАПОЛНЕНИЯ ШАБЛОНА: ', round(time.time()-start_time, 2), 'секунд')
                    print('ПРОЦЕСС: ' + str(process) + '%' + '\n ')
        print('ОБЩЕЕ ВРЕМЯ РАБОТЫ: ', round(time.time()-main_time, 2), 'секунд' + '\n ')
        
        print ('КОЛИЧЕСТВО КАТЕГОРИЙ В ФИДЕ:', category_list_len)
        print ('КОЛИЧЕСТВО ЗАПОЛНЕННЫХ ШАБЛОНОВ:', pro)
        
        if category_list_len!=pro:
            print ('ВНИМАНИЕ!!!' + '\n' + 'КОЛИЧЕСТВО ПРЕДОСТАВЛЕННЫХ КАТЕГОРИЙ В ФИДЕ НЕ СОВПАДАЕТ С КОЛИЧЕСТВОМ ЗАПОЛНЕННЫХ ШАБЛОНОВ!' +
                   '\n' + 'ЛИБО ПАПКЕ ИМЕЮТСЯ НЕ ВСЕ ШАБЛОНЫ ДЛЯ ЗАПОЛНЕНИЯ, ЛИБО В ФИДЕ УКАЗАНЫ ЛИШНИЕ КАТЕГОРИИ!')
    
#===============================================================================================================================================
#ЗАВЕРШЕНИЕ РАБОТЫ
    
    elif command=='Q' or command=='q':
        program_status=False
        print('ЗАВЕРШЕНИЕ РАБОТЫ')
    
#===============================================================================================================================================
#НЕВЕРНАЯ КОМАНДА
    
    else:
        print('ВЫБРАНА НЕВЕРНАЯ КОМАНДА\n')